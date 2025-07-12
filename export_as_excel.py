from openpyxl import Workbook
import io

def create_excel(match_data):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Home Team"
    ws1.append(["Name", "Goals", "Assists"])
    for player in match_data.get('home_players', []):
        ws1.append([player['name'], player['goals'], player['assists']])

    ws2 = wb.create_sheet("Away Team")
    ws2.append(["Name", "Goals", "Assists"])
    for player in match_data.get('away_players', []):
        ws2.append([player['name'], player['goals'], player['assists']])

    ws3 = wb.create_sheet("Match Info")
    ws3.append(["Home Team", match_data['home_team']])
    ws3.append(["Away Team", match_data['away_team']])
    ws3.append(["Score", match_data['match_score']])
    ws3.append(["Possession", match_data['possession']])
    ws3.append(["Man of the Match", match_data['motm']])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output